import csv
import logging
import os
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional
from fastapi import UploadFile
from app.core.exceptions.error_messages import ErrorKey
from app.core.exceptions.exception_classes import AppException
import re

logger = logging.getLogger(__name__)


@dataclass
class OCRConfig:
    enabled: bool = True
    dpi: int = 220
    lang: str = "eng"
    max_pages: Optional[int] = None


@dataclass
class ExtractorOptions:
    ocr: OCRConfig = field(default_factory=OCRConfig)
    enable_generic_textract_fallback: bool = True
    strict_pdf_header_check: bool = True

    # spreadsheet / delimited extraction guards
    excel_max_rows_per_sheet: Optional[int] = 10000   # None = no cap
    delimited_max_rows: Optional[int] = 100000
    include_sheet_headers: bool = True                # prepend "## Sheet: <name>"
    output_cell_sep: str = "\t"                       # normalize to TSV-like text
    output_row_sep: str = "\n"                        # row separator
    # how much to sample for dialect sniffing
    csv_sniff_limit_bytes: int = 4096

    # html
    # when not using Markdown, keep "text (URL)" for anchors
    html_preserve_links: bool = True
    html_block_separator: str = "\n"       # separator for extracted text blocks


class FileTextExtractor:
    """
        Unified text extractor for common document types.

        Inputs
        ------
        - path (str | Path): local filesystem path
        - filename and content (bytes): in-memory file; `filename` is used to infer type
        - file (fastapi.UploadFile): reads name and bytes from the upload

        Supported types
        ---------------
        - PDF (.pdf): pdfminer.six → OCR fallback (Tesseract/pdf2image) → textract fallback
        - Word: .docx via docx2txt; .doc via antiword/textract (when available)
        - Plain text: .txt, .md, .log (UTF-8 with latin-1 fallback)
        - Delimited: .csv, .tsv, .tab (dialect sniff; normalized to tab-separated text)
        - Excel: .xlsx (openpyxl), .xls (xlrd<2.0)
        - HTML: .html, .htm (BeautifulSoup/readability → plain text; optional Markdown if enabled)
        - Other/unknown: optional generic textract fallback

        Behavior
        --------
        - Returns extracted text as `str`.
        - Cleans up any temporary files it creates.
        - Raises AppException(ErrorKey.FILE_EXTRACT_USAGE) if called without a valid input combination.
        Examples
        --------
        extractor.extract(path="/data/report.pdf")
        extractor.extract(filename="notes.docx", content=blob)
        extractor.extract(file=upload_file)
    """

    # (moved .csv out to delimited handler)
    TEXT_SUFFIXES = {".txt", ".md", ".log"}
    DELIMITED_SUFFIXES = {".csv", ".tsv", ".tab"}
    # (extend later for .xlsb via pyxlsb if nedded)
    EXCEL_SUFFIXES = {".xlsx", ".xls"}
    HTML_SUFFIXES = {".html", ".htm"}

    def __init__(self, options: Optional[ExtractorOptions] = None):
        self.options = options or ExtractorOptions()

    # ---------- Public API ----------

    def extract(self, *, file: Optional[UploadFile] = None, filename: Optional[str] = None, content: Optional[bytes] = None,
                path: Optional[str | Path] = None) -> str:
        # Provide either (Path) or (filename and content) or file(UploadFile type).
        if path is not None:
            return self.extract_from_path(path)
        if filename is not None and content is not None:
            if len(content) == 0:
                logger.warning(
                    f"No content provided for {filename}, returning empty string")
                return ""
            return self.extract_from_bytes(filename, content)
        if file is not None:
            return self.extract_from_bytes(file.filename, file.file.read())
        raise AppException(ErrorKey.FILE_EXTRACT_USAGE)

    def extract_from_bytes(self, filename: str, content: bytes) -> str:
        suffix = (Path(filename).suffix or ".bin").lower()

        if suffix == ".pdf" and self.options.strict_pdf_header_check:
            if not self._looks_like_pdf(content[:8]):
                logger.warning(
                    f"[extractor] Not a real PDF for {filename} (head={content[:8]!r}); decoding as text.")
                return content.decode("utf-8", errors="replace")

        tmp_path: Optional[str] = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp.flush()
                tmp_path = tmp.name
            return self._extract_by_suffix(Path(tmp_path))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    def extract_from_path(self, path: str | Path) -> str:
        return self._extract_by_suffix(Path(path))

    # ---------- Routing ----------

    def _extract_by_suffix(self, path: Path) -> str:
        sfx = path.suffix.lower()

        if sfx == ".pdf":
            return self._extract_pdf(path)
        if sfx == ".docx":
            return self._extract_docx(path)
        if sfx == ".doc":
            return self._extract_doc(path)
        if sfx in self.EXCEL_SUFFIXES:
            return self._extract_excel(path)
        if sfx in self.DELIMITED_SUFFIXES:
            return self._extract_delimited(path)
        if sfx in self.HTML_SUFFIXES:
            return self._extract_html(path)
        if sfx in self.TEXT_SUFFIXES:
            return self._extract_plaintext(path)

        return self._extract_generic_textract(path) if self.options.enable_generic_textract_fallback else ""

    # ---------- Concrete extractors ----------

    def _extract_pdf(self, path: Path) -> str:
        txt = ""
        try:
            from pdfminer.high_level import extract_text as pdfminer_extract_text
            txt = pdfminer_extract_text(str(path)) or ""
            logger.info("[extractor] pdf used=pdfminer.six")
        except Exception as e:
            logger.info(f"[extractor] pdfminer failed: {e}")

        if self._is_mostly_pagebreaks(txt) and self.options.ocr.enabled:
            has_tesseract = bool(shutil.which("tesseract"))
            logger.info(
                f"[extractor] pdf empty; OCR fallback (tesseract={has_tesseract})")
            if has_tesseract:
                try:
                    ocr_txt = self._ocr_pdf(path)
                    if ocr_txt.strip():
                        logger.info("[extractor] pdf used=OCR(pytesseract)")
                        return ocr_txt
                except Exception as e:
                    logger.info(f"[extractor] OCR failed: {e}")
            try:
                import textract
                if shutil.which("pdftotext"):
                    txt = textract.process(str(path), method="pdftotext").decode(
                        "utf-8", errors="replace")
                    logger.info("[extractor] pdf used=textract(pdftotext)")
                else:
                    txt = textract.process(str(path), method="pdfminer").decode(
                        "utf-8", errors="replace")
                    logger.info("[extractor] pdf used=textract(pdfminer)")
            except Exception as e2:
                logger.info(f"[extractor] textract PDF fallbacks failed: {e2}")
        return txt

    def _extract_docx(self, path: Path) -> str:
        try:
            import docx2txt
            logger.info("[extractor] docx used=docx2txt")
            return docx2txt.process(str(path)) or ""
        except Exception as e:
            logger.info(f"[extractor] docx2txt failed: {e}")
            return self._extract_generic_textract(path)

    def _extract_plaintext(self, path: Path) -> str:
        logger.info("[extractor] used=plaintext")
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return path.read_text(encoding="latin-1", errors="replace")

    def _extract_doc(self, path: Path) -> str:
        # 1) Try textract with specific backends if present
        try:
            import textract

            if shutil.which("antiword"):
                logger.info("[extractor] doc used=textract(antiword)")
                return textract.process(str(path), method="antiword").decode("utf-8", errors="replace")
        except Exception as e:
            logger.info(f"[extractor] textract(.doc) failed: {e}")

        # 2) Last resort: generic textract (if enabled)
        return self._extract_generic_textract(path)

    def _extract_delimited(self, path: Path) -> str:
        """
        CSV/TSV/TAB → normalized tab-separated text.
        """
        logger.info("[extractor] used=csv(tsv/tab)")
        encodings = ("utf-8", "latin-1")
        sample_bytes = self.options.csv_sniff_limit_bytes

        for enc in encodings:
            try:
                with path.open("r", encoding=enc, errors="replace", newline="") as f:
                    sample = f.read(sample_bytes)
                    f.seek(0)

                    # Try sniffing; fall back to extension-based delimiter
                    delim = None
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                        delim = dialect.delimiter
                    except Exception:
                        sfx = path.suffix.lower()
                        if sfx == ".tsv" or sfx == ".tab":
                            delim = "\t"
                        else:
                            delim = ","

                    reader = csv.reader(f, delimiter=delim)
                    lines = []
                    max_rows = self.options.delimited_max_rows
                    for i, row in enumerate(reader):
                        if max_rows is not None and i >= max_rows:
                            break
                        lines.append(self.options.output_cell_sep.join(
                            self._fmt_cell(v) for v in row))
                    return self.options.output_row_sep.join(lines)
            except Exception as e:
                logger.info(
                    f"[extractor] CSV read with encoding={enc} failed: {e}")
        return ""

    def _extract_excel(self, path: Path) -> str:
        sfx = path.suffix.lower()
        out_parts: list[str] = []
        rows_cap = self.options.excel_max_rows_per_sheet

        if sfx == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=str(
                    path), read_only=True, data_only=True)
                for ws in wb.worksheets:
                    if self.options.include_sheet_headers:
                        out_parts.append(f"## Sheet: {ws.title}")
                    count = 0
                    for row in ws.iter_rows(values_only=True):
                        if rows_cap is not None and count >= rows_cap:
                            break
                        out_parts.append(self.options.output_cell_sep.join(
                            self._fmt_cell(v) for v in row))
                        count += 1
                logger.info("[extractor] xlsx used=openpyxl")
                return self.options.output_row_sep.join(out_parts)
            except Exception as e:
                logger.info(f"[extractor] openpyxl failed: {e}")
                # fallthrough → try xlrd if it happens to support (unlikely), else textract
        if sfx == ".xls":
            try:
                import xlrd  # xlrd<2.0 supports .xls
                book = xlrd.open_workbook(str(path))
                for sheet in book.sheets():
                    if self.options.include_sheet_headers:
                        out_parts.append(f"## Sheet: {sheet.name}")
                    count = 0
                    for r in range(sheet.nrows):
                        if rows_cap is not None and count >= rows_cap:
                            break
                        vals = []
                        for c in range(sheet.ncols):
                            vals.append(self._fmt_cell(sheet.cell_value(r, c)))
                        out_parts.append(
                            self.options.output_cell_sep.join(vals))
                        count += 1
                logger.info("[extractor] xls used=xlrd")
                return self.options.output_row_sep.join(out_parts)
            except Exception as e:
                logger.info(f"[extractor] xlrd failed: {e}")

        # Last resort for spreadsheets
        return self._extract_generic_textract(path)

    def _extract_html(self, path: Path) -> str:
        """
        HTML/HTM → clean text (or Markdown if enabled and available).
        Strategy:
          - Try readability-lxml to isolate main content (optional)
          - Parse with BeautifulSoup; drop scripts/styles/noscript
          - Optionally keep links as "text (URL)" or emit Markdown via html2text/markdownify
        """
        raw = path.read_bytes()

        # Try robust decoding first (BeautifulSoup's UnicodeDammit is great at this)
        try:
            from bs4 import UnicodeDammit

            decoded = UnicodeDammit(raw).unicode_markup or raw.decode(
                "utf-8", errors="replace")
        except Exception:
            decoded = raw.decode("utf-8", errors="replace")
        html_to_parse = decoded

        try:
            import html2text

            h = html2text.HTML2Text()
            h.ignore_links = not self.options.html_preserve_links
            h.ignore_images = True
            h.body_width = 0  # no hard wraps
            md = h.handle(html_to_parse or "")
            return self._normalize_whitespace(md)
        except Exception as e:
            logger.info(f"[extractor] html2text failed: {e}")

        # Plain-text extraction via BeautifulSoup
        try:
            from bs4 import BeautifulSoup

            # Prefer lxml if installed; fallback to html.parser
            parser = "lxml"
            try:
                import lxml  # noqa: F401
            except Exception:
                parser = "html.parser"

            soup = BeautifulSoup(html_to_parse or "", parser)

            # Remove non-text elements
            for tag in soup(["script", "style", "noscript", "template", "meta", "link", "iframe"]):
                tag.decompose()

            # Optionally preserve anchor targets (when not using Markdown)
            if self.options.html_preserve_links:
                for a in soup.find_all("a"):
                    href = (a.get("href") or "").strip()
                    text = a.get_text(strip=True)
                    # Replace link node with "text (URL)" or just URL if no text
                    if href:
                        repl = f"{text} ({href})" if text else href
                        a.replace_with(repl)
                    else:
                        a.replace_with(text)

            # Replace <br> with newlines to avoid run-on lines
            for br in soup.find_all("br"):
                br.replace_with(self.options.html_block_separator)

            # Get text with chosen block separator
            text = soup.get_text(separator=self.options.html_block_separator)

            return self._normalize_whitespace(text)
        except Exception as e:
            logger.info(f"[extractor] BeautifulSoup html parsing failed: {e}")
            # last resort: strip tags naively
            stripped = re.sub(r"<[^>]+>", " ", html_to_parse or "")
            return self._normalize_whitespace(stripped)

    # Small helper to collapse excessive whitespace/newlines

    def _normalize_whitespace(self, s: str) -> str:
        s = re.sub(r"[ \t]+\n", "\n", s)  # trailing spaces before newlines
        s = re.sub(r"\n{3,}", "\n\n", s)  # collapse 3+ blank lines to 2
        s = re.sub(r"[ \t]{2,}", " ", s)  # collapse runs of spaces/tabs
        return s.strip()

    def _extract_generic_textract(self, path: Path) -> str:
        if not self.options.enable_generic_textract_fallback:
            return ""
        try:
            import textract
            logger.info("[extractor] used=textract(generic)")
            return (textract.process(str(path)) or b"").decode("utf-8", errors="replace")
        except Exception as e:
            logger.info(f"[extractor] textract(generic) failed: {e}")
            return ""

    # ---------- Helpers ----------

    @staticmethod
    def _looks_like_pdf(head8: bytes) -> bool:
        return head8.startswith(b"%PDF-")

    @staticmethod
    def _is_mostly_pagebreaks(s: str) -> bool:
        return len(s.strip().replace("\x0c", "")) < 10

    def _ocr_pdf(self, pdf_path: Path) -> str:
        from pdf2image import convert_from_path
        import pytesseract
        imgs = convert_from_path(str(pdf_path), dpi=self.options.ocr.dpi)
        texts: list[str] = []
        for i, img in enumerate(imgs):
            if self.options.ocr.max_pages is not None and i >= self.options.ocr.max_pages:
                break
            texts.append(pytesseract.image_to_string(
                img, lang=self.options.ocr.lang))
        return "\n".join(texts)

    @staticmethod
    def _fmt_cell(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.isoformat()
        return str(v)


class FileExtractor:
    """Handles extraction of text content from various file formats"""

    @staticmethod
    def extract_from_pdf(file_path: str) -> str:
        """Extract text from PDF files"""
        try:
            text = FileTextExtractor().extract(path=file_path)
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting from PDF: {str(e)}")
            return ""

    @staticmethod
    def extract_from_docx(file_path: str) -> str:
        """Extract text from DOCX files"""
        try:
            text = FileTextExtractor().extract(path=file_path)
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting from DOCX: {str(e)}")
            return ""

    @staticmethod
    def extract_from_image(file_path: str) -> str:
        """Extract text from image files using OCR"""
        try:
            from PIL import Image
            import pytesseract
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting from image: {str(e)}")
            return ""

    @staticmethod
    def extract_from_txt(file_path: str) -> str:
        """Extract text from plain text files"""
        try:
            text = FileTextExtractor().extract(path=file_path)
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting from text file: {str(e)}")
            return ""
